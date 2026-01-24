from __future__ import annotations

import logging
from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo
from typing import Optional

from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, Message, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Config
from app.db.database import Database
from app.db import queries
from app.utils.keyboards import admin_panel_kb, main_menu_kb

logger = logging.getLogger(__name__)

router = Router()


def is_admin(user_id: int, config: Config) -> bool:
    """Проверяет, является ли пользователь админом"""
    return user_id in config.admin_ids


class BroadcastStates(StatesGroup):
    waiting_message = State()


@router.callback_query(F.data == "menu:admin")
async def admin_panel_handler(
    query: CallbackQuery,
    config: Config,
) -> None:
    """Обработчик кнопки админ-панели"""
    if query.from_user is None or query.message is None:
        return
    
    if not is_admin(query.from_user.id, config):
        await query.answer("❌ У вас нет доступа к админ-панели", show_alert=True)
        return
    
    await query.answer()
    await query.message.edit_text(
        "🔐 <b>Админ-панель</b>\n\n"
        "Выберите действие:",
        reply_markup=admin_panel_kb().as_markup(),
    )


@router.callback_query(F.data == "admin:back")
async def admin_back_handler(
    query: CallbackQuery,
    config: Config,
    db: Database,
) -> None:
    """Возврат в главное меню из админ-панели"""
    if query.from_user is None or query.message is None:
        return
    
    user = await queries.get_user_by_tg_id(db, query.from_user.id)
    if not user:
        await query.answer("❌ Пользователь не найден", show_alert=True)
        return
    
    await query.answer()
    await query.message.edit_text(
        "👋 <b>Главное меню</b>",
        reply_markup=main_menu_kb(config.admin_ids, query.from_user.id).as_markup(),
    )


@router.callback_query(F.data == "admin:stats")
async def admin_stats_handler(
    query: CallbackQuery,
    config: Config,
    db: Database,
    tz: ZoneInfo,
) -> None:
    """Подробная статистика бота"""
    if query.from_user is None or query.message is None:
        return
    
    if not is_admin(query.from_user.id, config):
        await query.answer("❌ У вас нет доступа", show_alert=True)
        return
    
    await query.answer("📊 Загрузка статистики...")
    
    try:
        # Получаем статистику
        users = await db.fetch_all("SELECT * FROM users")
        total_users = len(users)
        
        # Активные пользователи (за последние 30 дней)
        thirty_days_ago = datetime.now(tz) - timedelta(days=30)
        active_users = await db.fetch_all(
            "SELECT DISTINCT user_id FROM workout_logs WHERE date >= ?",
            (thirty_days_ago.isoformat(),)
        )
        active_count = len(active_users)
        
        # Новые пользователи за последние 7 дней
        seven_days_ago = datetime.now(tz) - timedelta(days=7)
        new_users = await db.fetch_all(
            "SELECT * FROM users WHERE created_at >= ?",
            (seven_days_ago.isoformat(),)
        )
        new_count = len(new_users)
        
        # Статистика тренировок
        total_workouts = await db.fetch_one(
            "SELECT COUNT(*) as count FROM workout_logs"
        )
        total_workouts_count = total_workouts["count"] if total_workouts else 0
        
        done_workouts = await db.fetch_one(
            "SELECT COUNT(*) as count FROM workout_logs WHERE status = 'done'"
        )
        done_count = done_workouts["count"] if done_workouts else 0
        
        missed_workouts = await db.fetch_one(
            "SELECT COUNT(*) as count FROM workout_logs WHERE status = 'missed'"
        )
        missed_count = missed_workouts["count"] if missed_workouts else 0
        
        # Статистика веса
        total_weight_entries = await db.fetch_one(
            "SELECT COUNT(*) as count FROM weights"
        )
        weight_count = total_weight_entries["count"] if total_weight_entries else 0
        
        # Пользователи с расписанием
        users_with_schedule = await db.fetch_one(
            "SELECT COUNT(DISTINCT user_id) as count FROM workout_schedule"
        )
        schedule_count = users_with_schedule["count"] if users_with_schedule else 0
        
        # Статистика за последние 7 дней
        workouts_last_7_days = await db.fetch_one(
            "SELECT COUNT(*) as count FROM workout_logs WHERE date >= ?",
            (seven_days_ago.isoformat(),)
        )
        workouts_7d = workouts_last_7_days["count"] if workouts_last_7_days else 0
        
        # Статистика за последние 30 дней
        workouts_last_30_days = await db.fetch_one(
            "SELECT COUNT(*) as count FROM workout_logs WHERE date >= ?",
            (thirty_days_ago.isoformat(),)
        )
        workouts_30d = workouts_last_30_days["count"] if workouts_last_30_days else 0
        
        # Процент выполнения тренировок
        completion_rate = (done_count / total_workouts_count * 100) if total_workouts_count > 0 else 0
        
        stats_text = (
            f"📊 <b>Статистика бота</b>\n\n"
            f"👥 <b>Пользователи:</b>\n"
            f"• Всего пользователей: <b>{total_users}</b>\n"
            f"• Активных (30 дней): <b>{active_count}</b>\n"
            f"• Новых (7 дней): <b>{new_count}</b>\n"
            f"• С расписанием: <b>{schedule_count}</b>\n\n"
            f"💪 <b>Тренировки:</b>\n"
            f"• Всего тренировок: <b>{total_workouts_count}</b>\n"
            f"• Выполнено: <b>{done_count}</b> ({completion_rate:.1f}%)\n"
            f"• Пропущено: <b>{missed_count}</b>\n"
            f"• За 7 дней: <b>{workouts_7d}</b>\n"
            f"• За 30 дней: <b>{workouts_30d}</b>\n\n"
            f"⚖️ <b>Вес:</b>\n"
            f"• Всего записей: <b>{weight_count}</b>\n\n"
            f"📅 <b>Обновлено:</b> {datetime.now(tz).strftime('%d.%m.%Y %H:%M')}"
        )
        
        await query.message.edit_text(
            stats_text,
            reply_markup=admin_panel_kb().as_markup(),
        )
    except Exception as e:
        logger.error(f"Ошибка при получении статистики: {e}", exc_info=True)
        await query.message.edit_text(
            "❌ <b>Ошибка при получении статистики</b>\n\n"
            f"Детали: {str(e)}",
            reply_markup=admin_panel_kb().as_markup(),
        )


@router.callback_query(F.data == "admin:export")
async def admin_export_handler(
    query: CallbackQuery,
    config: Config,
    db: Database,
    tz: ZoneInfo,
) -> None:
    """Выгрузка данных в Excel"""
    if query.from_user is None or query.message is None:
        return
    
    if not is_admin(query.from_user.id, config):
        await query.answer("❌ У вас нет доступа", show_alert=True)
        return
    
    await query.answer("📥 Создание Excel файла...")
    
    try:
        wb = Workbook()
        
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Лист 1: Пользователи
        ws_users = wb.create_sheet("Пользователи")
        ws_users.append(["ID", "Telegram ID", "Целевой вес", "Создан", "Смещение недели"])
        users = await db.fetch_all("SELECT * FROM users")
        for user_row in users:
            user = dict(user_row)
            ws_users.append([
                user["id"],
                user["tg_id"],
                user.get("target_weight") or "",
                user.get("created_at") or "",
                user.get("week_parity_offset") or 0,
            ])
        
        # Стилизация заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_users[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws_users.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_users.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Лист 2: Тренировки
        ws_workouts = wb.create_sheet("Тренировки")
        ws_workouts.append(["ID", "User ID", "Дата", "Статус", "Длительность", "Заметки"])
        workouts = await db.fetch_all("SELECT * FROM workout_logs ORDER BY date DESC")
        for workout_row in workouts:
            workout = dict(workout_row)
            ws_workouts.append([
                workout["id"],
                workout["user_id"],
                workout.get("date") or "",
                workout.get("status") or "",
                workout.get("duration") or "",
                workout.get("notes") or "",
            ])
        
        # Стилизация заголовков
        for cell in ws_workouts[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws_workouts.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_workouts.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Лист 3: Вес
        ws_weight = wb.create_sheet("Вес")
        ws_weight.append(["ID", "User ID", "Вес", "Дата"])
        weights = await db.fetch_all("SELECT * FROM weights ORDER BY date DESC")
        for weight_row in weights:
            weight_entry = dict(weight_row)
            ws_weight.append([
                weight_entry["id"],
                weight_entry["user_id"],
                weight_entry.get("weight") or "",
                weight_entry.get("date") or "",
            ])
        
        # Стилизация заголовков
        for cell in ws_weight[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws_weight.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_weight.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Лист 4: Расписание
        ws_schedule = wb.create_sheet("Расписание")
        ws_schedule.append(["ID", "User ID", "День недели", "Время", "Тип недели"])
        schedules = await db.fetch_all("SELECT * FROM workout_schedule")
        day_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        for schedule_row in schedules:
            schedule = dict(schedule_row)
            weekday = schedule.get("weekday", 0)
            ws_schedule.append([
                schedule["id"],
                schedule["user_id"],
                day_names[weekday] if weekday < len(day_names) else weekday,
                schedule.get("time") or "",
                schedule.get("week_type") or "any",
            ])
        
        # Стилизация заголовков
        for cell in ws_schedule[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws_schedule.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_schedule.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем в BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Читаем данные в bytes
        excel_data = excel_buffer.read()
        excel_buffer.close()
        
        # Отправляем файл
        filename = f"discipline_bot_export_{datetime.now(tz).strftime('%Y%m%d_%H%M%S')}.xlsx"
        file = BufferedInputFile(excel_data, filename=filename)
        
        await query.message.answer_document(
            document=file,
            caption=f"📥 <b>Выгрузка данных</b>\n\n"
                   f"📊 Листы: Пользователи, Тренировки, Вес, Расписание\n"
                   f"📅 Дата: {datetime.now(tz).strftime('%d.%m.%Y %H:%M')}",
        )
        
        await query.message.edit_text(
            "✅ <b>Файл успешно создан и отправлен!</b>",
            reply_markup=admin_panel_kb().as_markup(),
        )
    except Exception as e:
        logger.error(f"Ошибка при выгрузке в Excel: {e}", exc_info=True)
        await query.message.edit_text(
            "❌ <b>Ошибка при создании Excel файла</b>\n\n"
            f"Детали: {str(e)}",
            reply_markup=admin_panel_kb().as_markup(),
        )


@router.callback_query(F.data == "admin:users")
async def admin_users_handler(
    query: CallbackQuery,
    config: Config,
    db: Database,
    tz: ZoneInfo,
) -> None:
    """Список пользователей"""
    if query.from_user is None or query.message is None:
        return
    
    if not is_admin(query.from_user.id, config):
        await query.answer("❌ У вас нет доступа", show_alert=True)
        return
    
    await query.answer("👥 Загрузка списка пользователей...")
    
    try:
        users = await db.fetch_all("SELECT * FROM users ORDER BY created_at DESC LIMIT 50")
        
        if not users:
            await query.message.edit_text(
                "👥 <b>Список пользователей</b>\n\n"
                "Пользователей пока нет.",
                reply_markup=admin_panel_kb().as_markup(),
            )
            return
        
        users_text = "👥 <b>Последние 50 пользователей:</b>\n\n"
        for i, user_row in enumerate(users, 1):
            user = dict(user_row)
            created_at = user.get("created_at", "")
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    created_str = dt.astimezone(tz).strftime("%d.%m.%Y")
                except:
                    created_str = created_at[:10] if len(created_at) >= 10 else created_at
            else:
                created_str = "неизвестно"
            
            target_weight = user.get("target_weight") or "не указан"
            users_text += f"{i}. ID: {user['id']} | TG: {user['tg_id']} | Вес: {target_weight} | {created_str}\n"
        
        if len(users) == 50:
            users_text += "\n⚠️ Показаны только последние 50 пользователей"
        
        await query.message.edit_text(
            users_text,
            reply_markup=admin_panel_kb().as_markup(),
        )
    except Exception as e:
        logger.error(f"Ошибка при получении списка пользователей: {e}", exc_info=True)
        await query.message.edit_text(
            "❌ <b>Ошибка при получении списка пользователей</b>\n\n"
            f"Детали: {str(e)}",
            reply_markup=admin_panel_kb().as_markup(),
        )


@router.callback_query(F.data == "admin:broadcast")
async def admin_broadcast_start(
    query: CallbackQuery,
    config: Config,
    state: FSMContext,
) -> None:
    """Начало рассылки сообщений"""
    if query.from_user is None or query.message is None:
        return
    
    if not is_admin(query.from_user.id, config):
        await query.answer("❌ У вас нет доступа", show_alert=True)
        return
    
    await query.answer()
    await state.set_state(BroadcastStates.waiting_message)
    await query.message.edit_text(
        "📢 <b>Рассылка сообщений</b>\n\n"
        "Введите сообщение для рассылки всем пользователям.\n\n"
        "Используйте HTML разметку для форматирования.\n"
        "Для отмены отправьте /cancel",
        reply_markup=None,
    )


@router.message(BroadcastStates.waiting_message)
async def admin_broadcast_send(
    message: Message,
    config: Config,
    db: Database,
    state: FSMContext,
) -> None:
    """Отправка рассылки"""
    if message.from_user is None or message.text is None:
        return
    
    if not is_admin(message.from_user.id, config):
        await message.answer("❌ У вас нет доступа")
        await state.clear()
        return
    
    if message.text.startswith("/cancel"):
        await state.clear()
        await message.answer("❌ Рассылка отменена", reply_markup=admin_panel_kb().as_markup())
        return
    
    await message.answer("📤 Начинаю рассылку...")
    
    try:
        users = await db.fetch_all("SELECT tg_id FROM users")
        total = len(users)
        success = 0
        failed = 0
        
        for user in users:
            try:
                await message.bot.send_message(
                    user["tg_id"],
                    message.text,
                    parse_mode="HTML",
                )
                success += 1
            except Exception as e:
                logger.warning(f"Не удалось отправить сообщение пользователю {user['tg_id']}: {e}")
                failed += 1
        
        await state.clear()
        await message.answer(
            f"✅ <b>Рассылка завершена!</b>\n\n"
            f"📊 Статистика:\n"
            f"• Всего пользователей: {total}\n"
            f"• Успешно отправлено: {success}\n"
            f"• Ошибок: {failed}",
            reply_markup=admin_panel_kb().as_markup(),
        )
    except Exception as e:
        logger.error(f"Ошибка при рассылке: {e}", exc_info=True)
        await state.clear()
        await message.answer(
            f"❌ <b>Ошибка при рассылке</b>\n\n"
            f"Детали: {str(e)}",
            reply_markup=admin_panel_kb().as_markup(),
        )
